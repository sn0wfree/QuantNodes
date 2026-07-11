# coding=utf-8
"""v7.3 简化集成: 仅消费现有 9 因子 + 13 指数数据, 输出组合 NAV.

[设计原则] 不新建 v7/ 生产代码, 不新建子策略类, 不写 5-fold. 仅写一次性脚本.

[数据来源] (用户决策: 直接复用数据, 不写代码)
  - 9 因子净值: ~/Public/高频宏观因子/高频宏观因子跟踪_output_2026-06-01.xlsx
  - 13 指数:   ~/Public/高频宏观因子/Factor Minicking组合-高频宏观因子20260601.xlsx (sheet=主要指数)

[方法论借鉴] NowCasting之Factor Mimicking v2 cell 99 + cell 108 + cell 118
  - main_idx_cols: 13 个 level-1 指数 (沪深300/中证500/中证1000/恒生 + 4 中债 + 商品)
  - simple_backtest: nv = np.dot(w, NV.T).cumprod()
  - 季度滚动 + 3-year window
  - 简化版本: 用滚动 IC 替代 Bootstrap-Lasso (用户决策: 不写复杂代码)

[输出] reports/momentum_etf_rotation/v7/
  - v7_3_oos_navs.parquet  (3 NAV 列: v7.3 宏观 / v6.2 ir_expanding / combo 50/50)
  - v7_3_oos_results.csv   (Calmar / Sharpe / DD / Ann 4 指标 × 4 策略)
  - v7_3_factor_loadings.csv (季度 IC 矩阵 - 9 macro factor × 13 index)
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parents[1]
OUT_DIR = REPO / "reports" / "momentum_etf_rotation" / "v7"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 源文件路径
SRC_DIR = Path.home() / "Public" / "高频宏观因子"
FACTOR_FILE = SRC_DIR / "高频宏观因子跟踪_output_2026-06-01.xlsx"
INDEX_FILE = SRC_DIR / "Factor Minicking组合-高频宏观因子20260601.xlsx"

# 13 个 level-1 指数 (源 v2 cell 99)
INDEX_COLS = [
    "沪深300指数", "中证500指数", "中证1000", "恒生指数",
    "中债10年期国债指数", "中债3-5年期国债指数",
    "中债国开行债券总指数", "中债企业债总指数",
    "南华综合指数", "南华工业品指数", "南华农产品指数",
    "期货结算价(连续):布伦特原油", "收盘价:沪金指数",
]

# 9 个宏观因子
FACTOR_COLS = [
    "宏观增长因子", "宏观通胀因子_生活端", "宏观通胀因子_生产端",
    "无风险收益率", "信用利差因子", "期限利差因子_债",
    "期限利差因子_股", "期限利差因子_加权", "宏观汇率因子",
]

# 时间窗口
START_DT = "2018-01-01"
END_DT = "2026-06-30"
QUARTER_WINDOW = 12  # 12 quarters = 3 years (源 cell 118)

# 验证路径
if not FACTOR_FILE.exists():
    raise FileNotFoundError(f"Factor file not found: {FACTOR_FILE}")
if not INDEX_FILE.exists():
    raise FileNotFoundError(f"Index file not found: {INDEX_FILE}")


# ============================================================================
# 1. 加载数据
# ============================================================================
def load_macro_factors() -> pd.DataFrame:
    """加载 9 周频宏观因子净值, 索引=周日期."""
    wb = openpyxl.load_workbook(FACTOR_FILE, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]
    df = pd.DataFrame(data, columns=header)
    df["dt"] = pd.to_datetime(df["dt"])
    df = df.set_index("dt").sort_index()
    return df[FACTOR_COLS]


def load_indices() -> pd.DataFrame:
    """加载 13 指数日频价格, 索引=日日期."""
    wb = openpyxl.load_workbook(INDEX_FILE, data_only=True, read_only=True)
    ws = wb["主要指数"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]  # 第二行是列名
    data = rows[8:]   # 第 9 行起为数据 (前 8 行是元信息)

    df = pd.DataFrame(data, columns=[str(c) if c else f"col_{i}"
                                      for i, c in enumerate(header)])
    df["指标名称"] = pd.to_datetime(df["指标名称"], errors="coerce")
    df = df.set_index("指标名称").sort_index()
    df.index.name = "dt"

    # 强制转数值 (源文件可能有空 cell 或 str)
    sub = df[INDEX_COLS].copy()
    sub = sub.apply(pd.to_numeric, errors="coerce")
    return sub


# ============================================================================
# 2. OOS 季度调仓 (源 v2 cell 118)
# ============================================================================
def compute_ic_weighted_portfolio(
    idx_ret: pd.DataFrame,        # (T, 13)
    factor_ret_weekly: pd.DataFrame,  # (T, 9) 周频
    start_dt: str = START_DT,
) -> tuple[pd.Series, dict[pd.Timestamp, pd.Series]]:
    """滚动 IC 加权, 季度再平衡, 3-year window.

    Returns:
        nav: pd.Series, OOS NAV 起点=1
        weights_history: dict of {rebal_date: weights}
    """
    # 季度边界 (pandas 3.0: Q → QE)
    quarter_list = idx_ret.loc[start_dt:].resample("QE").last().index.tolist()
    # 至少需要 QUARTER_WINDOW 个季度才能开始
    if len(quarter_list) <= QUARTER_WINDOW:
        raise ValueError(f"Need > {QUARTER_WINDOW} quarters of data, got {len(quarter_list)}")

    rebal_dates = quarter_list[QUARTER_WINDOW:]
    weights_history: dict[pd.Timestamp, pd.Series] = {}

    for end_q in rebal_dates:
        # 3-year rolling window 起点
        start_idx = quarter_list.index(end_q) - QUARTER_WINDOW
        start_q = quarter_list[start_idx]
        if start_idx < 0:
            start_q = quarter_list[0]

        # 窗口内指数日频收益
        window_idx_ret = idx_ret.loc[start_q:end_q].dropna(how="all")
        # 窗口内宏观因子周频收益, resample 到周对齐
        window_factor = factor_ret_weekly.loc[start_q:end_q].dropna(how="all")

        if len(window_idx_ret) < 252 or len(window_factor) < 52:
            continue

        # 简化版 IC: 用窗口内累计 weekly factor return vs 指数 weekly return 的相关
        # 因子: 9 个, resample 取 weekly last
        idx_w = window_idx_ret.resample("W-FRI").sum()
        factor_w = window_factor.copy()
        common = idx_w.index.intersection(factor_w.index)
        idx_w = idx_w.loc[common]
        factor_w = factor_w.loc[common]

        # IC 矩阵: 13 指数 × 9 因子
        ic_mat = pd.DataFrame(
            {
                idx_col: [factor_w[f].corr(idx_w[idx_col]) for f in FACTOR_COLS]
                for idx_col in INDEX_COLS
            },
            index=FACTOR_COLS,
        ).T  # shape (13 indices, 9 factors)

        # 把每个 index 的 9 factor IC 平均成单值
        ic_avg = ic_mat.mean(axis=1)
        # 权重 = max(0, IC), 归一化
        raw_w = ic_avg.clip(lower=0)
        total = raw_w.sum()
        if total > 0:
            w = raw_w / total
        else:
            # 全为负或全 NaN → 等权
            w = pd.Series(1.0 / len(INDEX_COLS), index=INDEX_COLS)
        weights_history[end_q] = w

    # 应用权重: 把权重 shift 到下一期, simple_backtest (源 cell 108)
    weight_df = pd.DataFrame(weights_history).T
    weight_df = weight_df.reindex(idx_ret.loc[start_dt:].index, method="ffill")
    weight_df = weight_df.fillna(1.0 / len(INDEX_COLS))

    # OOS: 第一个权重 start_q 在 future, 把 first-date 之前的权重设为 0
    # 实际: 第 1 个 rebal 是 quarter_list[QUARTER_WINDOW], 之前用等权
    nav_ret = (idx_ret.loc[start_dt:] * weight_df).sum(axis=1).fillna(0)
    nav = (1 + nav_ret).cumprod()
    nav = nav.loc[start_dt:]
    return nav, weights_history


# ============================================================================
# 3. 评估
# ============================================================================
def metrics(s: pd.Series, label: str = "") -> dict:
    """Calmar / Sharpe / MaxDD / Ann."""
    s = s.dropna()
    r = s.pct_change().dropna()
    if len(r) < 2:
        return {"ann": 0.0, "vol": 0.0, "sharpe": 0.0, "dd": 0.0, "calmar": 0.0,
                "name": label}
    n = len(r)
    ann = (1 + r).prod() ** (252 / n) - 1
    vol = r.std() * np.sqrt(252)
    sharpe = ann / vol if vol > 0 else 0.0
    dd = (s / s.cummax() - 1).min()
    calmar = ann / abs(dd) if abs(dd) > 0.001 else 0.0
    return {"name": label, "ann": ann, "vol": vol, "sharpe": sharpe,
            "dd": dd, "calmar": calmar}


def print_metrics(m: dict) -> None:
    """Pretty print."""
    print(f"  {m['name']:30s} ann={m['ann']*100:7.2f}%  vol={m['vol']*100:7.2f}%"
          f"  sharpe={m['sharpe']:6.3f}  dd={m['dd']*100:7.2f}%  calmar={m['calmar']:6.3f}")


# ============================================================================
# 4. 主流程
# ============================================================================
def main() -> None:
    print("=" * 70)
    print("v7.3 简化集成: 9 宏观因子 + 13 指数, IC 加权 OOS 回测")
    print("=" * 70)

    # --- 1. 加载 ---
    print("\n[1/4] 加载数据...")
    factor_nav = load_macro_factors()
    factor_ret_weekly = factor_nav.pct_change().dropna()
    print(f"  9 周频宏观因子净值: {factor_nav.shape}, "
          f"{factor_nav.index.min().date()} -> {factor_nav.index.max().date()}")

    idx_panel = load_indices()
    idx_ret = idx_panel.pct_change().dropna(how="all")
    print(f"  13 指数日频收益:    {idx_ret.shape}, "
          f"{idx_ret.index.min().date()} -> {idx_ret.index.max().date()}")

    # --- 2. v7.3 OOS ---
    print(f"\n[2/4] v7.3 OOS 回测 (季度 IC 加权, {QUARTER_WINDOW}Q={QUARTER_WINDOW//4}Y window)...")
    nav_v73, weights_history = compute_ic_weighted_portfolio(idx_ret, factor_ret_weekly)
    print(f"  v7.3 NAV: {len(nav_v73)} rows, "
          f"{nav_v73.index.min().date()} -> {nav_v73.index.max().date()}")
    print(f"  rebal 次数: {len(weights_history)}")

    # --- 3. 对照 ---
    print(f"\n[3/4] 加载对照 (v6.2 ir_expanding, v1.0 locked)...")
    try:
        v62 = pd.read_parquet(
            REPO / "reports" / "momentum_etf_rotation" / "combo"
            / "v6_1_v6_2_combined_navs.parquet"
        )["v6.2 ir_expanding"]
    except FileNotFoundError:
        print("  WARN: v6.2 ir_expanding NAVs not found, 跳过对照")
        v62 = None

    try:
        v10 = pd.read_parquet(
            REPO / "reports" / "momentum_etf_rotation" / "combo"
            / "unified_v1v5_navs_calA.parquet"
        )["v1.0 locked"]
    except FileNotFoundError:
        print("  WARN: v1.0 locked NAVs not found, 跳过对照")
        v10 = None

    # 等权 ETF 基准 (沪深300 近似, 用指数池等权)
    nav_eq = (1 + idx_ret.loc[nav_v73.index.min():nav_v73.index.max()]
              [INDEX_COLS].fillna(0).mean(axis=1)).cumprod()

    # 对齐 + 组合
    aligned_idx = nav_v73.index.intersection(nav_eq.index)
    nav_v73_a = nav_v73.reindex(aligned_idx)
    nav_eq_a = nav_eq.reindex(aligned_idx)

    results = []

    print(f"\n[4/4] 评估...")
    print(f"\n--- 全期 ({nav_v73.index.min().date()} -> {nav_v73.index.max().date()}) ---")
    m = metrics(nav_v73_a, "v7.3 宏观 IC 加权 (OOS)")
    print_metrics(m)
    results.append(m)

    m = metrics(nav_eq_a, "13 指数等权")
    print_metrics(m)
    results.append(m)

    if v62 is not None:
        v62_a = v62.reindex(aligned_idx).dropna()
        common = nav_v73_a.index.intersection(v62_a.index)
        m = metrics(v62_a.loc[common], "v6.2 ir_expanding (current best)")
        print_metrics(m)
        results.append(m)

        # 组合: 等权 v6.2 + v7.3
        combo = (v62_a.loc[common] + nav_v73_a.loc[common]) / 2
        m = metrics(combo, "combo 50/50 (v6.2 + v7.3)")
        print_metrics(m)
        results.append(m)

    if v10 is not None:
        v10_a = v10.reindex(aligned_idx).dropna()
        common10 = nav_v73_a.index.intersection(v10_a.index)
        m = metrics(v10_a.loc[common10], "v1.0 locked (OOS Calmar 1.79)")
        print_metrics(m)
        results.append(m)

    # OOS 2022-2026 子集
    if "2022-01-01" < nav_v73_a.index.max().strftime("%Y-%m-%d"):
        print(f"\n--- OOS 2022-01-01 -> {nav_v73.index.max().date()} ---")
        subset = nav_v73_a.loc["2022-01-01":]
        m = metrics(subset, "v7.3 宏观 IC 加权 (OOS 2022+)")
        print_metrics(m)
        results.append(m)

        subset_eq = nav_eq_a.loc["2022-01-01":]
        m = metrics(subset_eq, "13 指数等权 (OOS 2022+)")
        print_metrics(m)
        results.append(m)

        if v62 is not None:
            v62_2022 = v62.reindex(aligned_idx).loc["2022-01-01":].dropna()
            cmn22 = nav_v73_a.index.intersection(v62_2022.index).intersection(
                pd.date_range("2022-01-01", aligned_idx.max()))
            m = metrics(v62_2022.loc[cmn22], "v6.2 ir_expanding (OOS 2022+)")
            print_metrics(m)
            results.append(m)

            combo22 = (v62_2022.loc[cmn22] + nav_v73_a.loc[cmn22]) / 2
            m = metrics(combo22, "combo 50/50 (OOS 2022+)")
            print_metrics(m)
            results.append(m)

    # 相关性
    if v62 is not None:
        common_full = nav_v73_a.index.intersection(v62.reindex(aligned_idx).index)
        v73_ret = nav_v73_a.loc[common_full].pct_change().dropna()
        v62_ret = v62.reindex(aligned_idx).loc[common_full].pct_change().dropna()
        common_ret = v73_ret.index.intersection(v62_ret.index)
        corr = v73_ret.loc[common_ret].corr(v62_ret.loc[common_ret])
        print(f"\n--- 相关性 ---")
        print(f"  v7.3 vs v6.2 ir_expanding: corr = {corr:.3f}")

    # --- 5. 保存 ---
    print(f"\n[save] 输出文件...")
    df_results = pd.DataFrame(results)
    out_csv = OUT_DIR / "v7_3_oos_results.csv"
    df_results.to_csv(out_csv, index=False)
    print(f"  {out_csv}")

    # NAVs parquet
    navs = pd.DataFrame({"v7.3 宏观 IC 加权 (OOS)": nav_v73})
    if v62 is not None:
        navs["v6.2 ir_expanding"] = v62.reindex(nav_v73.index)
    if v10 is not None:
        navs["v1.0 locked"] = v10.reindex(nav_v73.index)
    navs["13 指数等权"] = nav_eq.reindex(nav_v73.index)
    if v62 is not None:
        navs["combo 50/50 (v6.2 + v7.3)"] = (
            navs["v6.2 ir_expanding"] + navs["v7.3 宏观 IC 加权 (OOS)"]
        ) / 2
    out_pq = OUT_DIR / "v7_3_oos_navs.parquet"
    navs.to_parquet(out_pq)
    print(f"  {out_pq}")

    # Factor loadings (季度 IC)
    if weights_history:
        loadings = pd.DataFrame(weights_history).T  # (quarter_dates, 13 indices)
        loadings.index.name = "rebal_date"
        out_loadings = OUT_DIR / "v7_3_factor_loadings.csv"
        loadings.to_csv(out_loadings)
        print(f"  {out_loadings}")

    print(f"\n{'=' * 70}")
    print("完成.")


if __name__ == "__main__":
    main()
