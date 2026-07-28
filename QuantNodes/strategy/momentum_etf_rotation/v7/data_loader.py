# coding=utf-8
"""v7 数据加载: 统一返回价格/净值, 收益由调用方计算.

数据源:
  - 8 周频宏观因子净值: ~/Public/高频宏观因子/高频宏观因子跟踪_output_2026-06-01.xlsx
  - 13 指数日价格:       ~/Public/高频宏观因子/Factor Minicking组合-高频宏观因子20260601.xlsx
  - 51 ETF 日 NAV:       data/real/etf_nav_*.parquet

核心函数:
  load_aligned_prices(pool, start) -> dict(asset_prices, factor_nav, benchmark)

原则: 数据导入只返回价格/净值, 不返回收益. 收益计算在策略层完成.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parents[4]
HF_DIR = REPO / "data" / "high_freq_macro"
HF_DIR.mkdir(parents=True, exist_ok=True)
REAL_DIR = REPO / "data" / "real"

SRC_DIR = Path.home() / "Public" / "高频宏观因子"
FACTOR_FILE = SRC_DIR / "高频宏观因子跟踪_output_2026-06-01.xlsx"
INDEX_FILE = SRC_DIR / "Factor Minicking组合-高频宏观因子20260601.xlsx"

# 8 宏观因子 (移除 期限利差因子_加权，无价值)
FACTOR_COLS = [
    "宏观增长因子", "宏观通胀因子_生活端", "宏观通胀因子_生产端",
    "无风险收益率", "信用利差因子", "期限利差因子_债",
    "期限利差因子_股", "宏观汇率因子",
]

# 13 INDICES
INDEX_COLS = [
    "沪深300指数",
    "中证500指数",
    "中证1000",
    "恒生指数",
    "中债10年期国债指数",
    "中债3-5年期国债指数",
    "中债1-3年国债财富指数",
    "中债国开行债券总指数",
    "中债企业债总指数",
    "南华工业品指数",
    "南华农产品指数",
    "期货结算价(连续):布伦特原油",
    "收盘价:沪金指数",
]

# ── EXPANDED POOL: 51 ETFs + 5 bond indices = 56 assets ──────────────────

EXPANDED_ETF_COLS: list[str] = [
    "510300", "510500", "510050", "159915", "588000", "159901",
    "512760", "512480", "515030", "515790", "512690", "512170", "512010",
    "515050", "159928", "512880", "512000", "512800", "515220", "512200",
    "512400", "512660", "512980", "515880", "159996", "512120",
    "510900", "159920", "513010", "513050", "159740",
    "518880", "518800", "159985", "161226", "159981", "159766",
    "513100", "513300", "513500", "513520", "513880", "159941",
    "510880", "512890", "512260", "515900", "512040", "159786", "515080", "515100",
]

EXPANDED_BOND_INDICES: list[str] = [
    "中债10年期国债指数",
    "中债3-5年期国债指数",
    "中债1-3年国债财富指数",
    "中债国开行债券总指数",
    "中债企业债总指数",
]

EXPANDED_COLS: list[str] = EXPANDED_ETF_COLS + EXPANDED_BOND_INDICES

EQUITY_ETF_COLS: list[str] = [
    "510300", "510500", "510050", "159915", "588000", "159901",
    "512760", "512480", "515030", "515790", "512690", "512170", "512010",
    "515050", "159928", "512880", "512000", "512800", "515220", "512200",
    "512400", "512660", "512980", "515880", "159996", "512120",
    "510900", "159920", "513010", "513050", "159740",
    "513100", "513300", "513500", "513520", "513880", "159941",
    "510880", "512890", "512260", "515900", "512040", "159786", "515080", "515100",
]

COMMODITY_ETF_COLS: list[str] = [
    "518880", "518800", "159985", "161226", "159981", "159766",
]


# ============================================================
# 底层加载: 因子净值
# ============================================================
def load_macro_factors() -> pd.DataFrame:
    """加载 8 周频宏观因子净值, 索引=周日期."""
    cache = HF_DIR / "v9_factors_weekly.parquet"
    if cache.exists():
        return pd.read_parquet(cache)
    if not FACTOR_FILE.exists():
        raise FileNotFoundError(f"Factor file not found: {FACTOR_FILE}")
    wb = openpyxl.load_workbook(FACTOR_FILE, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]
    df = pd.DataFrame(data, columns=header)
    df["dt"] = pd.to_datetime(df["dt"])
    df = df.set_index("dt").sort_index()
    out = df[FACTOR_COLS].astype(float)
    out.to_parquet(cache)
    return out


# ============================================================
# 底层加载: benchmark 价格
# ============================================================
def load_benchmark_price(benchmark: str = "沪深300指数") -> pd.Series:
    """加载 benchmark 指数日价格 (用于 trend filter 信号)."""
    cache = HF_DIR / f"v9_benchmark_{benchmark.replace('指数','')}.parquet"
    if cache.exists():
        return pd.read_parquet(cache).squeeze("columns")

    prices = _load_index_prices_from_excel()
    if benchmark not in prices.columns:
        raise ValueError(f"benchmark {benchmark!r} not in 主要指数 sheet")
    s = prices[benchmark].copy()
    bday_idx = pd.bdate_range(start=s.dropna().index[0], end=s.index[-1])
    s = s.reindex(bday_idx).bfill()
    s.name = benchmark
    s.to_frame().to_parquet(cache)
    return s


# ============================================================
# 底层加载: 13 指数日价格 (从 Excel)
# ============================================================
def _load_index_prices_from_excel() -> pd.DataFrame:
    """从 Excel 读取 13 指数日价格 (未做 bday 重采样).

    Returns:
        DataFrame (T, 13) 升序, 含 INDEX_COLS.
    """
    if not INDEX_FILE.exists():
        raise FileNotFoundError(f"Index file not found: {INDEX_FILE}")
    wb = openpyxl.load_workbook(INDEX_FILE, data_only=True, read_only=True)

    # ── 主要指数 sheet: 12 列 ──
    ws = wb["主要指数"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    data = rows[8:]
    df = pd.DataFrame(data, columns=[str(c) if c else f"col_{i}"
                                      for i, c in enumerate(header)])
    df["指标名称"] = pd.to_datetime(df["指标名称"], errors="coerce")
    df = df.dropna(subset=["指标名称"])
    df = df.set_index("指标名称").sort_index()
    df.index.name = "dt"

    main_cols = [c for c in INDEX_COLS if c in df.columns]
    sub = df[main_cols].apply(pd.to_numeric, errors="coerce")

    # ── 指数 sheet: 1-3 年国债财富指数 ──
    if "中债1-3年国债财富指数" in INDEX_COLS and "中债1-3年国债财富指数" not in sub.columns:
        ws_idx = wb["指数"]
        idx_rows = list(ws_idx.iter_rows(values_only=True))
        idx_header = idx_rows[0]
        col_name = "1-3 年国债财富指数"
        if col_name in idx_header:
            idx_df = pd.DataFrame(
                idx_rows[1:],
                columns=[str(c) if c else f"col_{i}" for i, c in enumerate(idx_header)],
            )
            idx_df["dt"] = pd.to_datetime(idx_df["dt"], errors="coerce")
            idx_df = idx_df.dropna(subset=["dt"])
            idx_df = idx_df[~idx_df["dt"].duplicated(keep="first")]
            idx_df = idx_df.set_index("dt").sort_index()
            idx_series = idx_df[col_name].apply(pd.to_numeric, errors="coerce")
            sub["中债1-3年国债财富指数"] = idx_series

    return sub


def _load_index_prices(start: str = "2008-01-01") -> pd.DataFrame:
    """加载 13 指数日价格, bday 重采样 + bfill.

    Returns:
        DataFrame (T_daily, 13) 含 INDEX_COLS 日价格.
    """
    cache = HF_DIR / "v7_index_prices.parquet"
    if cache.exists():
        return pd.read_parquet(cache).loc[start:]

    sub = _load_index_prices_from_excel()
    bday_idx = pd.bdate_range(start=sub.dropna().index[0], end=sub.index[-1])
    sub = sub.reindex(bday_idx).bfill()
    sub = sub[INDEX_COLS]
    sub.index.name = "dt"
    sub.to_parquet(cache)
    return sub.loc[start:]


# ============================================================
# 底层加载: 51 ETF 日 NAV (从 parquet)
# ============================================================
def _load_etf_nav() -> pd.DataFrame:
    """加载 51 ETF 日 NAV (合并 main + smartbeta)."""
    main_nav = pd.read_parquet(REAL_DIR / "etf_nav_2018-01-01_2026-06-30.parquet")
    smart_nav = pd.read_parquet(REAL_DIR / "etf_nav_smartbeta_2018-01-01_2026-06-30.parquet")
    smart_unique = [c for c in smart_nav.columns if c not in main_nav.columns]
    etf_nav = pd.concat([main_nav, smart_nav[smart_unique]], axis=1)
    return etf_nav[EXPANDED_ETF_COLS]


def _load_expanded_prices(start: str = "2018-01-01") -> pd.DataFrame:
    """加载 56 资产日价格 (51 ETF NAV + 5 债券指数价格).

    ETF 上市前保持 NaN (不 fill).
    债券指数 bfill 跨节假日缺口.

    Returns:
        DataFrame (T_daily, 56) 含 EXPANDED_COLS 日价格/NAV.
    """
    cache = HF_DIR / "v7_expanded_prices.parquet"
    if cache.exists():
        return pd.read_parquet(cache).loc[start:]

    # ETF NAV
    etf_nav = _load_etf_nav()

    # 债券指数价格
    idx_prices = _load_index_prices()
    bond_prices = idx_prices[EXPANDED_BOND_INDICES]

    # 对齐到共同 bday 范围
    common_start = max(etf_nav.dropna(how="all").index[0],
                       bond_prices.dropna(how="all").index[0])
    common_end = min(etf_nav.index[-1], bond_prices.index[-1])
    bday_idx = pd.bdate_range(start=common_start, end=common_end)

    etf_nav = etf_nav.reindex(bday_idx)
    bond_prices = bond_prices.reindex(bday_idx).bfill()

    expanded = pd.concat([etf_nav, bond_prices], axis=1)
    expanded = expanded[EXPANDED_COLS]
    expanded.index.name = "dt"
    expanded.to_parquet(cache)
    return expanded.loc[start:]


# ============================================================
# 公共接口: 统一加载对齐数据
# ============================================================
def load_aligned_prices(
    pool: str = "index",
    start: str = "2008-01-01",
) -> dict[str, pd.DataFrame | pd.Series]:
    """加载 v7.3 策略所需全部数据 (价格/净值, 不含收益).

    Parameters:
        pool: "index" (13 指数) 或 "expanded" (51 ETF + 5 债券 = 56)
        start: 起始日期

    Returns:
        {
            "asset_prices": DataFrame (T_daily, N),  日频价格/NAV
            "factor_nav":   DataFrame (T_weekly, 8), 周频因子净值
            "benchmark":    Series (T_daily,),       日频沪深300价格
        }
    """
    if pool == "index":
        asset_prices = _load_index_prices(start=start)
    elif pool == "expanded":
        asset_prices = _load_expanded_prices(start=start)
    else:
        raise ValueError(f"pool must be 'index' or 'expanded', got {pool!r}")

    factor_nav = load_macro_factors()
    benchmark = load_benchmark_price("沪深300指数")

    return {
        "asset_prices": asset_prices,
        "factor_nav": factor_nav,
        "benchmark": benchmark,
    }


__all__ = [
    "FACTOR_COLS",
    "INDEX_COLS",
    "EXPANDED_COLS",
    "EXPANDED_ETF_COLS",
    "EXPANDED_BOND_INDICES",
    "EQUITY_ETF_COLS",
    "COMMODITY_ETF_COLS",
    "load_macro_factors",
    "load_benchmark_price",
    "load_aligned_prices",
    "HF_DIR",
]
